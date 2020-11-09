import openpyxl
import os

'''workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
# print(str(sheet['A1'].value))
# print(str(sheet['B1'].value))

print(sheet.cell(row=1, column=3).value)

for i in range (1, 8):
    print(i, sheet.cell(row=i, column=2).value)
'''

wb = openpyxl.Workbook()
sheets = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheets[0])

sheet['A1'] = 42
sheet['A2'] = 'Hello'
sheet2 = wb.create_sheet()
sheet2.title = 'New sheet name'
wb.save('example.xlsx')

wb.create_sheet(index=0, title='My Other Sheet')
